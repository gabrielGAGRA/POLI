"""
Simulação e Visualização Técnica de Fonte Regulada com Diodo Zener
===================================================================
Laboratório de Eletrônica - POLI / USP
Circuito: Trafo (110:18) + Ponte Retificadora + Filtro C + Zener 1N4735 (6.20 V)
Exportação Técnica: Gráficos de Alta Resolução e Planilha Excel com Fórmulas e Gráficos Nativos
"""

import os
import sys
from dataclasses import dataclass
from typing import Tuple, Optional

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# =============================================================================
# 1. MODELAGEM FÍSICA E PARÂMETROS DO CIRCUITO
# =============================================================================

@dataclass(frozen=True)
class CircuitParams:
    """Parâmetros físicos e elétricos da fonte regulada."""
    f_rede: float = 60.0          # Frequência da rede CA (Hz)
    v_in_ef: float = 125.0        # Tensão eficaz primário (V)
    n_prim: float = 110.0         # Espiras primário
    n_sec: float = 18.0           # Espiras secundário
    v_diode: float = 0.8          # Queda direta por diodo (V)
    v_app: float = 0.407          # Tensão de ripple pico a pico no nó A (V)
    v_zener: float = 6.20         # Tensão de regulação do Zener (V)

    @property
    def t_rede(self) -> float:
        """Período da rede CA (s)."""
        return 1.0 / self.f_rede

    @property
    def f_ripple(self) -> float:
        """Frequência de ondulação retificada (Hz)."""
        return 2.0 * self.f_rede

    @property
    def t_ripple(self) -> float:
        """Período de ondulação retificada (s)."""
        return 1.0 / self.f_ripple

    @property
    def e0_ef(self) -> float:
        """Tensão eficaz no secundário do transformador (V)."""
        return self.v_in_ef * (self.n_sec / self.n_prim)

    @property
    def e0_pk(self) -> float:
        """Tensão de pico no secundário (V)."""
        return self.e0_ef * np.sqrt(2.0)

    @property
    def va_max(self) -> float:
        """Tensão de pico / máxima no nó A (V)."""
        return self.e0_pk - 2.0 * self.v_diode

    @property
    def va_min(self) -> float:
        """Tensão mínima no nó A com carga (V)."""
        return self.va_max - self.v_app

    @property
    def va_dc(self) -> float:
        """Tensão média contínua no nó A (V)."""
        return 0.5 * (self.va_max + self.va_min)


def simulate_waveforms(
    params: CircuitParams,
    t_max_ms: float = 33.333333,
    num_points: int = 6000
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """
    Gera as formas de onda no domínio do tempo para o circuito da fonte regulada.

    Retorna:
        t_ms: Vetor de tempo em milissegundos
        v_rect: Tensão senoidal retificada em onda completa (|e0(t)| - 2*VD)
        v_a: Tensão filtrada no capacitor (Nó A) com regime de carga e descarga
        v_s: Tensão contínua regulada pelo Diodo Zener (Nó B)
    """
    t_s = np.linspace(0, t_max_ms * 1e-3, num_points)
    t_ms = t_s * 1e3

    omega = 2.0 * np.pi * params.f_rede
    v_sec = params.e0_pk * np.sin(omega * t_s)

    # Onda senoidal retificada em ponte completa (2 quedas de diodo)
    v_rect_raw = np.abs(v_sec) - 2.0 * params.v_diode
    v_rect = np.maximum(0.0, v_rect_raw)

    # Instante de condução e pico angular
    sin_cond = (params.va_min + 2.0 * params.v_diode) / params.e0_pk
    sin_cond = np.clip(sin_cond, -1.0, 1.0)
    theta_cond = np.arcsin(sin_cond)
    t_cond_rel = theta_cond / omega
    t_pk_rel = (np.pi / 2.0) / omega

    delta_t_charge = t_pk_rel - t_cond_rel
    delta_t_discharge = params.t_ripple - delta_t_charge

    # Constante de tempo tau para reproduzir com exatidão o ripple VApp
    tau_disch = delta_t_discharge / np.log(params.va_max / params.va_min)

    v_a = np.zeros_like(t_s)
    half_period = params.t_ripple

    for i, t in enumerate(t_s):
        cycle_idx = int(np.floor(t / half_period))
        t_local = t - cycle_idx * half_period

        if t_local < t_cond_rel:
            t_from_prev_peak = t_local + (half_period - t_pk_rel)
            v_a[i] = params.va_max * np.exp(-t_from_prev_peak / tau_disch)
        elif t_local <= t_pk_rel:
            v_a[i] = v_rect[i]
        else:
            t_from_peak = t_local - t_pk_rel
            v_a[i] = params.va_max * np.exp(-t_from_peak / tau_disch)

    # Tensão regulada de saída no Nó B (Vs = Vz)
    v_s = np.full_like(t_s, params.v_zener)

    return t_ms, v_rect, v_a, v_s


# =============================================================================
# 2. VISUALIZAÇÃO GRÁFICA TÉCNICA (MATPLOTLIB)
# =============================================================================

def plot_circuit_waveforms(
    params: CircuitParams,
    save_path: str = "formas_de_onda_fonte_excel.png",
    show_plot: bool = False
) -> plt.Figure:
    """
    Gera o gráfico técnico com 3 subplots alinhados temporalmente:
      1. Nó A - Retificação e Filtragem (Visão Geral 0-30 V)
      2. Nó A - Detalhe da Ondulação / Ripple com Cotas (26.7 - 27.6 V)
      3. Nó B - Saída Regulada pelo Diodo Zener 1N4735 (0-10 V)
    """
    t_ms, v_rect, v_a, v_s = simulate_waveforms(params)

    plt.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Segoe UI", "DejaVu Sans", "Helvetica", "Arial"],
        "mathtext.fontset": "stixsans",
        "axes.edgecolor": "#334155",
        "axes.linewidth": 1.1,
        "grid.color": "#E2E8F0",
        "grid.linestyle": "--",
        "grid.linewidth": 0.8,
        "grid.alpha": 0.85,
    })

    fig, (ax1, ax2, ax3) = plt.subplots(
        nrows=3,
        ncols=1,
        figsize=(13.0, 10.5),
        sharex=True,
        gridspec_kw={"height_ratios": [1.15, 1.25, 0.9], "hspace": 0.22}
    )
    fig.patch.set_facecolor("#FFFFFF")

    # Paleta de cores corporativa / de engenharia
    c_sine = "#64748B"
    c_va = "#1D4ED8"
    c_vs = "#059669"
    c_grid_vert = "#94A3B8"
    c_red = "#DC2626"
    c_dark_red = "#991B1B"

    time_ticks = [0.0, 8.33, 16.67, 25.00, 33.33]

    # =========================================================================
    # SUBPLOT 1: NÓ A - VISÃO GERAL (RETIFICAÇÃO E FILTRAGEM)
    # =========================================================================
    ax1.set_facecolor("#F8FAFC")
    ax1.grid(True, which="both", zorder=0)

    ax1.plot(
        t_ms, v_rect,
        color=c_sine,
        linestyle="--",
        linewidth=1.7,
        alpha=0.75,
        label=r"Onda Senoidal Retificada em Ponte: $v_{rect}(t) = |e_0(t)| - 2V_D$",
        zorder=2
    )
    ax1.plot(
        t_ms, v_a,
        color=c_va,
        linewidth=2.5,
        label=r"Tensão no Filtro Capacitivo: $v_A(t)$ (Carga)",
        zorder=4
    )
    ax1.axhline(params.va_max, color=c_red, linestyle=":", linewidth=1.0, alpha=0.8, zorder=1)

    ax1.annotate(
        rf"$V_{{A,\max}} = {params.va_max:.2f}\ \mathrm{{V}}$",
        xy=(4.167, params.va_max),
        xytext=(4.167, 30.5),
        arrowprops=dict(arrowstyle="->", color=c_red, lw=1.2),
        ha="center", fontsize=9.0, fontweight="bold", color="#B91C1C",
        bbox=dict(boxstyle="round,pad=0.2", facecolor="#FEF2F2", edgecolor="#FCA5A5", alpha=0.95),
        zorder=6
    )
    ax1.annotate(
        rf"$E_{{0,pk}} - 2V_D = {params.va_max:.2f}\ \mathrm{{V}}$",
        xy=(20.833, params.va_max),
        xytext=(20.833, 30.5),
        arrowprops=dict(arrowstyle="->", color="#475569", lw=1.2),
        ha="center", fontsize=8.5, color="#334155",
        bbox=dict(boxstyle="round,pad=0.2", facecolor="#F1F5F9", edgecolor="#CBD5E1", alpha=0.95),
        zorder=6
    )

    ax1.set_ylabel(r"Tensão $v_A$ (V)", fontsize=10.5, fontweight="bold", labelpad=8)
    ax1.set_title(
        r"(a) Estágio de Retificação e Filtragem Capacitiva (Nó A - Visão Geral): $f_{ripple} = 120\ \mathrm{Hz}$",
        fontsize=11.5, fontweight="bold", pad=8, color="#0F172A"
    )
    ax1.set_ylim(-1.5, 33.5)
    ax1.legend(loc="lower right", frameon=True, facecolor="#FFFFFF", framealpha=0.95, edgecolor="#CBD5E1", fontsize=8.5)

    # =========================================================================
    # SUBPLOT 2: NÓ A - DETALHAMENTO DO RIPPLE E COTAS NUMÉRICAS
    # =========================================================================
    ax2.set_facecolor("#F8FAFC")
    ax2.grid(True, which="both", zorder=0)

    ax2.plot(
        t_ms, v_rect,
        color=c_sine,
        linestyle="--",
        linewidth=1.7,
        alpha=0.75,
        label=r"$v_{rect}(t)$ (crista)",
        zorder=2
    )
    ax2.plot(
        t_ms, v_a,
        color=c_va,
        linewidth=2.6,
        label=r"$v_A(t)$ no Nó A",
        zorder=4
    )

    ax2.axhline(params.va_max, color=c_red, linestyle=":", linewidth=1.2, alpha=0.9, zorder=1)
    ax2.axhline(params.va_dc,  color="#475569", linestyle="-.", linewidth=1.1, alpha=0.85, zorder=1)
    ax2.axhline(params.va_min, color=c_red, linestyle=":", linewidth=1.2, alpha=0.9, zorder=1)

    ax2.text(
        0.3, params.va_max + 0.03,
        rf"$V_{{A,\max}} = {params.va_max:.2f}\ \mathrm{{V}}$",
        color="#B91C1C", fontsize=9.0, fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.2", facecolor="#FEF2F2", edgecolor="#FCA5A5", alpha=0.95),
        zorder=6
    )
    ax2.text(
        0.3, params.va_dc + 0.03,
        rf"$V_{{A,\mathrm{{dc}}}} = {params.va_dc:.2f}\ \mathrm{{V}}$ (Médio)",
        color="#334155", fontsize=8.5, fontweight="semibold",
        bbox=dict(boxstyle="round,pad=0.2", facecolor="#F1F5F9", edgecolor="#CBD5E1", alpha=0.95),
        zorder=6
    )
    ax2.text(
        0.3, params.va_min - 0.08,
        rf"$V_{{A,\min}} = {params.va_min:.2f}\ \mathrm{{V}}$",
        color="#B91C1C", fontsize=9.0, fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.2", facecolor="#FEF2F2", edgecolor="#FCA5A5", alpha=0.95),
        zorder=6
    )

    t_cota = 20.833
    ax2.annotate(
        "",
        xy=(t_cota, params.va_min),
        xytext=(t_cota, params.va_max),
        arrowprops=dict(arrowstyle="<->", color=c_red, lw=1.8, shrinkA=0, shrinkB=0),
        zorder=6
    )
    ax2.text(
        t_cota + 0.4, (params.va_max + params.va_min) / 2.0,
        rf"$V_{{App}} = {params.v_app:.3f}\ \mathrm{{V}}\ (0.41\ \mathrm{{V}})$" + "\n" + r"(Ripple Pico a Pico)",
        color=c_dark_red, fontsize=8.5, fontweight="bold", va="center",
        bbox=dict(boxstyle="round,pad=0.22", facecolor="#FFFBEB", edgecolor="#FCD34D", alpha=0.95),
        zorder=6
    )

    ax2.annotate(
        "Recarga rápida\n(Condução dos diodos)",
        xy=(12.35, 27.20),
        xytext=(9.8, 27.42),
        arrowprops=dict(arrowstyle="->", color=c_va, lw=1.3, connectionstyle="arc3,rad=-0.15"),
        fontsize=8.5, color=c_va, fontweight="semibold", ha="center",
        bbox=dict(boxstyle="round,pad=0.2", facecolor="#EFF6FF", edgecolor="#BFDBFE", alpha=0.95),
        zorder=6
    )
    ax2.annotate(
        "Descarga lenta no capacitor\n(Diodos reversamente polarizados)",
        xy=(16.5, 27.05),
        xytext=(16.5, 26.78),
        arrowprops=dict(arrowstyle="->", color="#334155", lw=1.3, connectionstyle="arc3,rad=0.15"),
        fontsize=8.5, color="#334155", ha="center",
        bbox=dict(boxstyle="round,pad=0.2", facecolor="#F8FAFC", edgecolor="#CBD5E1", alpha=0.95),
        zorder=6
    )

    ax2.set_ylabel(r"Tensão $v_A$ (V)", fontsize=10.5, fontweight="bold", labelpad=8)
    ax2.set_title(
        r"(b) Detalhamento da Ondulação de Tensão (Ripple no Filtro Capacitivo - Nó A)",
        fontsize=11.5, fontweight="bold", pad=8, color="#0F172A"
    )
    ax2.set_ylim(26.70, 27.58)
    ax2.yaxis.set_major_formatter(ticker.FormatStrFormatter("%.2f"))
    ax2.legend(loc="upper right", frameon=True, facecolor="#FFFFFF", framealpha=0.95, edgecolor="#CBD5E1", fontsize=8.5)

    # =========================================================================
    # SUBPLOT 3: NÓ B - SAÍDA REGULADA PELO DIODO ZENER (Vs)
    # =========================================================================
    ax3.set_facecolor("#F8FAFC")
    ax3.grid(True, which="both", zorder=0)

    ax3.plot(
        t_ms, v_s,
        color=c_vs,
        linewidth=2.6,
        label=rf"Tensão Regulada de Saída: $V_s = V_Z = {params.v_zener:.2f}\ \mathrm{{V}}$ (1N4735)",
        zorder=3
    )
    ax3.axhline(params.v_zener, color="#047857", linestyle="--", linewidth=1.2, alpha=0.7, zorder=1)

    ax3.text(
        0.4, params.v_zener + 0.65,
        rf"$V_s = {params.v_zener:.2f}\ \mathrm{{V}}$ (Tensão Contínua Estável / DC)",
        color="#065F46", fontsize=9.0, fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.22", facecolor="#ECFDF5", edgecolor="#A7F3D0", alpha=0.95),
        zorder=5
    )
    ax3.annotate(
        r"Regulação Zener em condução reversa: Ripple residual $\Delta V_s \approx 0\ \mathrm{V}$",
        xy=(16.67, params.v_zener),
        xytext=(16.67, params.v_zener - 2.5),
        arrowprops=dict(arrowstyle="->", color=c_vs, lw=1.3),
        fontsize=8.5, color="#065F46", fontweight="semibold",
        ha="center",
        bbox=dict(boxstyle="round,pad=0.22", facecolor="#ECFDF5", edgecolor="#6EE7B7", alpha=0.95),
        zorder=6
    )

    ax3.set_ylabel(r"Tensão $V_s$ (V)", fontsize=10.5, fontweight="bold", labelpad=8)
    ax3.set_title(
        r"(c) Estágio de Regulação de Tensão (Nó B / Carga $R_L$ - Zener 1N4735)",
        fontsize=11.5, fontweight="bold", pad=8, color="#0F172A"
    )
    ax3.set_ylim(0.0, 10.0)
    ax3.legend(loc="upper right", frameon=True, facecolor="#FFFFFF", framealpha=0.95, edgecolor="#CBD5E1", fontsize=8.5)

    # Eixo X comum
    ax3.set_xlabel(r"Tempo $t$ (ms)", fontsize=11, fontweight="bold", labelpad=8)
    ax3.set_xlim(0.0, 33.333333)

    time_labels = [
        r"$\mathbf{0\ ms}$",
        r"$\mathbf{8.33\ ms}$" + "\n" + r"($T_{ripple}$)",
        r"$\mathbf{16.67\ ms}$" + "\n" + r"($T_{rede}$)",
        r"$\mathbf{25.00\ ms}$" + "\n" + r"($3T_{rip}$)",
        r"$\mathbf{33.33\ ms}$" + "\n" + r"($2T_{rede}$)"
    ]

    for t_val in time_ticks:
        ax1.axvline(t_val, color=c_grid_vert, linestyle=":", linewidth=1.0, alpha=0.75, zorder=1)
        ax2.axvline(t_val, color=c_grid_vert, linestyle=":", linewidth=1.0, alpha=0.75, zorder=1)
        ax3.axvline(t_val, color=c_grid_vert, linestyle=":", linewidth=1.0, alpha=0.75, zorder=1)

    ax3.set_xticks(time_ticks)
    ax3.set_xticklabels(time_labels, fontsize=9.0)

    fig.suptitle(
        "Formas de Onda de Tensão da Fonte Regulada (Retificador em Ponte, Filtro C e Diodo Zener)",
        fontsize=13.5, fontweight="bold", y=0.985, color="#0F172A"
    )

    fig.savefig(save_path, dpi=350, bbox_inches="tight")
    print(f"[OK] Gráfico técnico salvo com sucesso em: {save_path}")

    # Também salva no nome padrão para compatibilidade de histórico
    if save_path != "formas_de_onda_fonte.png":
        try:
            fig.savefig("formas_de_onda_fonte.png", dpi=350, bbox_inches="tight")
        except Exception:
            pass

    if show_plot:
        plt.show()

    return fig


# =============================================================================
# 3. GERAÇÃO DA PLANILHA EXCEL TÉCNICA (OPENPYXL)
# =============================================================================

def generate_excel_workbook(
    params: CircuitParams,
    excel_path: str = "formas_de_onda_fonte_excel.xlsx",
    img_path: Optional[str] = "formas_de_onda_fonte_excel.png",
    num_excel_points: int = 1000
) -> str:
    """
    Cria uma pasta de trabalho profissional do Microsoft Excel (.xlsx) estruturada em 3 abas:
      1. 'Resumo e Parâmetros': Tabela de parâmetros com fórmulas dinâmicas do Excel e gráfico embutido.
      2. 'Dados da Simulação': Dados tabulados no domínio do tempo para as 3 formas de onda.
      3. 'Gráficos Excel': Gráficos nativos do Excel (ScatterChart) com traçado contínuo.

    Retorna o caminho absoluto do arquivo salvo.
    """
    wb = openpyxl.Workbook()

    # Definição de estilos corporativos e de engenharia
    fnt_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    fnt_sub = Font(name="Segoe UI", size=10, italic=True, color="475569")
    fnt_section = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fnt_tbl_header = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    fnt_body = Font(name="Segoe UI", size=10, color="0F172A")
    fnt_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    fnt_math = Font(name="Consolas", size=9, italic=True, color="334155")

    fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_blue_sec = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fill_header_tbl = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_row_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_row_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    thin_border_side = Side(style="thin", color="CBD5E1")
    tbl_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # =========================================================================
    # ABA 1: RESUMO E PARÂMETROS
    # =========================================================================
    ws_params = wb.active
    ws_params.title = "Resumo e Parâmetros"
    ws_params.views.sheetView[0].showGridLines = True

    # Banner de Cabeçalho
    ws_params.merge_cells("B2:E2")
    ws_params["B2"] = "LABORATÓRIO DE ELETRÔNICA - POLI / USP"
    ws_params["B2"].font = fnt_title
    ws_params["B2"].fill = fill_navy
    ws_params["B2"].alignment = align_center
    ws_params.row_dimensions[2].height = 28

    ws_params.merge_cells("B3:E3")
    ws_params["B3"] = "Modelagem e Simulação da Fonte Regulada com Retificador em Ponte e Zener (1N4735)"
    ws_params["B3"].font = fnt_sub
    ws_params["B3"].alignment = align_center
    ws_params.row_dimensions[3].height = 18

    # Seção 1: Parâmetros de Entrada
    ws_params.merge_cells("B5:E5")
    ws_params["B5"] = "1. PARÂMETROS FÍSICOS DE ENTRADA DO CIRCUITO"
    ws_params["B5"].font = fnt_section
    ws_params["B5"].fill = fill_blue_sec
    ws_params["B5"].alignment = align_left
    ws_params.row_dimensions[5].height = 22

    headers_in = ["Parâmetro", "Valor", "Unidade", "Descrição Física"]
    for col_idx, text in enumerate(headers_in, start=2):
        cell = ws_params.cell(row=6, column=col_idx, value=text)
        cell.font = fnt_tbl_header
        cell.fill = fill_header_tbl
        cell.alignment = align_center if col_idx in (3, 4) else align_left
        cell.border = tbl_border
    ws_params.row_dimensions[6].height = 20

    input_data = [
        ("Frequência da Rede (f)", params.f_rede, "Hz", "Frequência nominal da rede CA brasileira", "0.0"),
        ("Tensão Eficaz no Primário (Vin)", params.v_in_ef, "Vef", "Tensão alternada eficaz na tomada", "0.0"),
        ("Espiras do Primário (N1)", params.n_prim, "espiras", "Número de espiras no primário do trafo", "0"),
        ("Espiras do Secundário (N2)", params.n_sec, "espiras", "Número de espiras no secundário do trafo", "0"),
        ("Queda Direta por Diodo (VD)", params.v_diode, "V", "Tensão de polarização direta por diodo da ponte", "0.00"),
        ("Ripple Pico a Pico no Nó A (VApp)", params.v_app, "V", "Ondulação de pico a pico no filtro capacitivo", "0.000"),
        ("Tensão Zener Nominal (Vz)", params.v_zener, "V", "Tensão estabilizada de regulação do 1N4735", "0.00"),
    ]

    for row_idx, (name, val, unit, desc, num_fmt) in enumerate(input_data, start=7):
        r_fill = fill_row_even if row_idx % 2 == 0 else None
        ws_params.row_dimensions[row_idx].height = 19

        c_name = ws_params.cell(row=row_idx, column=2, value=name)
        c_name.font = fnt_body
        c_name.border = tbl_border
        c_name.alignment = align_left
        if r_fill:
            c_name.fill = r_fill

        c_val = ws_params.cell(row=row_idx, column=3, value=val)
        c_val.font = fnt_bold
        c_val.border = tbl_border
        c_val.alignment = align_right
        c_val.number_format = num_fmt
        if r_fill:
            c_val.fill = r_fill

        c_unit = ws_params.cell(row=row_idx, column=4, value=unit)
        c_unit.font = fnt_body
        c_unit.border = tbl_border
        c_unit.alignment = align_center
        if r_fill:
            c_unit.fill = r_fill

        c_desc = ws_params.cell(row=row_idx, column=5, value=desc)
        c_desc.font = fnt_sub
        c_desc.border = tbl_border
        c_desc.alignment = align_left
        if r_fill:
            c_desc.fill = r_fill

    # Seção 2: Grandezas Físicas Calculadas com Fórmulas do Excel
    ws_params.merge_cells("B15:E15")
    ws_params["B15"] = "2. GRANDEZAS CALCULADAS VIA FÓRMULAS DO EXCEL"
    ws_params["B15"].font = fnt_section
    ws_params["B15"].fill = fill_navy
    ws_params["B15"].alignment = align_left
    ws_params.row_dimensions[15].height = 22

    headers_calc = ["Grandeza Física", "Valor / Fórmula Excel", "Unidade", "Equação / Relação Teórica"]
    for col_idx, text in enumerate(headers_calc, start=2):
        cell = ws_params.cell(row=16, column=col_idx, value=text)
        cell.font = fnt_tbl_header
        cell.fill = fill_header_tbl
        cell.alignment = align_center if col_idx in (3, 4) else align_left
        cell.border = tbl_border
    ws_params.row_dimensions[16].height = 20

    calc_rows = [
        ("Período da Rede Elétrica (T_rede)", "=(1/C7)*1000", "ms", "= (1 / f) * 1000", "0.00"),
        ("Frequência de Ondulação (f_ripple)", "=2*C7", "Hz", "= 2 * f (Ponte Completa)", "0"),
        ("Período de Ondulação (T_ripple)", "=(1/C18)*1000", "ms", "= (1 / f_ripple) * 1000", "0.00"),
        ("Tensão Eficaz Secundário (E0_ef)", "=C8*(C10/C9)", "Vef", "= Vin * (N2 / N1)", "0.000"),
        ("Tensão de Pico Secundário (E0_pk)", "=C20*SQRT(2)", "V", "= E0_ef * SQRT(2)", "0.000"),
        ("Tensão Máxima no Filtro (VA_max)", "=C21-2*C11", "V", "= E0_pk - 2 * VD", "0.00"),
        ("Tensão Mínima no Filtro (VA_min)", "=C22-C12", "V", "= VA_max - VApp", "0.00"),
        ("Tensão Média Contínua (VA_dc)", "=(C22+C23)/2", "V", "= (VA_max + VA_min) / 2", "0.00"),
        ("Tensão Regulada de Saída (Vs)", "=C13", "V", "= Vz (Regulador Zener)", "0.00"),
    ]

    for row_idx, (name, formula_val, unit, math_eq, num_fmt) in enumerate(calc_rows, start=17):
        is_highlight = "VA_max" in name or "VA_min" in name or "Vs" in name
        r_fill = fill_row_highlight if is_highlight else (fill_row_even if row_idx % 2 == 0 else None)
        ws_params.row_dimensions[row_idx].height = 19

        c_name = ws_params.cell(row=row_idx, column=2, value=name)
        c_name.font = fnt_bold if is_highlight else fnt_body
        c_name.border = tbl_border
        c_name.alignment = align_left
        if r_fill:
            c_name.fill = r_fill

        c_val = ws_params.cell(row=row_idx, column=3, value=formula_val)
        c_val.font = fnt_bold
        c_val.border = tbl_border
        c_val.alignment = align_right
        c_val.number_format = num_fmt
        if r_fill:
            c_val.fill = r_fill

        c_unit = ws_params.cell(row=row_idx, column=4, value=unit)
        c_unit.font = fnt_body
        c_unit.border = tbl_border
        c_unit.alignment = align_center
        if r_fill:
            c_unit.fill = r_fill

        c_math = ws_params.cell(row=row_idx, column=5, value=math_eq)
        c_math.font = fnt_math
        c_math.border = tbl_border
        c_math.alignment = align_left
        if r_fill:
            c_math.fill = r_fill

    # Ajuste automático das larguras das colunas A-E
    ws_params.column_dimensions["A"].width = 3.0
    ws_params.column_dimensions["B"].width = 38.0
    ws_params.column_dimensions["C"].width = 24.0
    ws_params.column_dimensions["D"].width = 12.0
    ws_params.column_dimensions["E"].width = 34.0
    ws_params.column_dimensions["F"].width = 4.0

    # Inserção da imagem técnica de alta resolução ao lado das tabelas (Coluna G)
    if img_path and os.path.exists(img_path):
        try:
            xl_img = XLImage(img_path)
            xl_img.width = 660
            xl_img.height = 530
            ws_params.add_image(xl_img, "G2")
        except Exception as e:
            print(f"[Aviso] Não foi possível embutir a imagem na planilha: {e}")

    # =========================================================================
    # ABA 2: DADOS DA SIMULAÇÃO (SÉRIE TEMPORAL)
    # =========================================================================
    ws_data = wb.create_sheet(title="Dados da Simulação")
    ws_data.views.sheetView[0].showGridLines = True

    t_ms, v_rect, v_a, v_s = simulate_waveforms(params, t_max_ms=33.333333, num_points=num_excel_points)

    data_headers = [
        "Tempo t (ms)",
        "v_rect(t) (V) - Senoide Retificada",
        "v_A(t) (V) - Filtro Capacitivo",
        "v_S(t) (V) - Saída Zener 1N4735"
    ]

    for col_idx, header in enumerate(data_headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = tbl_border
    ws_data.row_dimensions[1].height = 24

    for i in range(len(t_ms)):
        row_num = i + 2
        r_fill = fill_row_even if i % 2 == 0 else None

        c_t = ws_data.cell(row=row_num, column=1, value=float(t_ms[i]))
        c_t.number_format = "0.000"
        c_t.font = fnt_body
        c_t.alignment = align_right
        c_t.border = tbl_border

        c_rec = ws_data.cell(row=row_num, column=2, value=float(v_rect[i]))
        c_rec.number_format = "0.00"
        c_rec.font = fnt_body
        c_rec.alignment = align_right
        c_rec.border = tbl_border

        c_va = ws_data.cell(row=row_num, column=3, value=float(v_a[i]))
        c_va.number_format = "0.00"
        c_va.font = fnt_bold
        c_va.alignment = align_right
        c_va.border = tbl_border

        c_vs = ws_data.cell(row=row_num, column=4, value=float(v_s[i]))
        c_vs.number_format = "0.00"
        c_vs.font = fnt_bold
        c_vs.alignment = align_right
        c_vs.border = tbl_border

        if r_fill:
            c_t.fill = r_fill
            c_rec.fill = r_fill
            c_va.fill = r_fill
            c_vs.fill = r_fill

    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = f"A1:D{len(t_ms) + 1}"

    ws_data.column_dimensions["A"].width = 16.0
    ws_data.column_dimensions["B"].width = 32.0
    ws_data.column_dimensions["C"].width = 30.0
    ws_data.column_dimensions["D"].width = 30.0

    # =========================================================================
    # ABA 3: GRÁFICOS EXCEL NATIVOS (SCATTER CHARTS COM LINHAS SUAVES)
    # =========================================================================
    ws_charts = wb.create_sheet(title="Gráficos Excel")
    ws_charts.views.sheetView[0].showGridLines = True

    n_rows = len(t_ms) + 1
    xvalues = Reference(ws_data, min_col=1, min_row=2, max_row=n_rows)

    # Gráfico 1: Retificação e Filtragem (Nó A - Visão Geral 0-32 V)
    chart1 = ScatterChart()
    chart1.title = "Estágio de Retificação e Filtragem Capacitiva (Nó A - Visão Geral)"
    chart1.style = 13
    chart1.x_axis.title = "Tempo t (ms)"
    chart1.y_axis.title = "Tensão (V)"
    chart1.x_axis.scaling.min = 0.0
    chart1.x_axis.scaling.max = 33.33
    chart1.y_axis.scaling.min = 0.0
    chart1.y_axis.scaling.max = 32.0
    chart1.width = 24
    chart1.height = 12

    y_rect = Reference(ws_data, min_col=2, min_row=1, max_row=n_rows)
    s_rect = Series(y_rect, xvalues, title_from_data=True)
    s_rect.marker.symbol = None
    s_rect.graphicalProperties.line.solidFill = "64748B"
    s_rect.graphicalProperties.line.dashStyle = "dash"
    chart1.series.append(s_rect)

    y_va = Reference(ws_data, min_col=3, min_row=1, max_row=n_rows)
    s_va = Series(y_va, xvalues, title_from_data=True)
    s_va.marker.symbol = None
    s_va.graphicalProperties.line.solidFill = "1D4ED8"
    s_va.graphicalProperties.line.width = 25000
    chart1.series.append(s_va)

    ws_charts.add_chart(chart1, "B2")

    # Gráfico 2: Detalhamento do Ripple (Nó A - Zoom da Crista 26.50 - 27.60 V)
    chart2 = ScatterChart()
    chart2.title = "Detalhamento da Ondulação de Tensão (Ripple no Filtro Capacitivo - Nó A)"
    chart2.style = 13
    chart2.x_axis.title = "Tempo t (ms)"
    chart2.y_axis.title = "Tensão vA (V)"
    chart2.x_axis.scaling.min = 0.0
    chart2.x_axis.scaling.max = 33.33
    chart2.y_axis.scaling.min = 26.50
    chart2.y_axis.scaling.max = 27.60
    chart2.width = 24
    chart2.height = 12

    s_rect2 = Series(y_rect, xvalues, title_from_data=True)
    s_rect2.marker.symbol = None
    s_rect2.graphicalProperties.line.solidFill = "64748B"
    s_rect2.graphicalProperties.line.dashStyle = "dash"
    chart2.series.append(s_rect2)

    s_va2 = Series(y_va, xvalues, title_from_data=True)
    s_va2.marker.symbol = None
    s_va2.graphicalProperties.line.solidFill = "1D4ED8"
    s_va2.graphicalProperties.line.width = 25000
    chart2.series.append(s_va2)

    ws_charts.add_chart(chart2, "B26")

    # Gráfico 3: Estágio de Regulação Zener (Nó B - 0 a 10 V)
    chart3 = ScatterChart()
    chart3.title = "Estágio de Regulação de Tensão (Nó B - Diodo Zener 1N4735)"
    chart3.style = 13
    chart3.x_axis.title = "Tempo t (ms)"
    chart3.y_axis.title = "Tensão de Saída Vs (V)"
    chart3.x_axis.scaling.min = 0.0
    chart3.x_axis.scaling.max = 33.33
    chart3.y_axis.scaling.min = 0.0
    chart3.y_axis.scaling.max = 10.0
    chart3.width = 24
    chart3.height = 12

    y_vs = Reference(ws_data, min_col=4, min_row=1, max_row=n_rows)
    s_vs = Series(y_vs, xvalues, title_from_data=True)
    s_vs.marker.symbol = None
    s_vs.graphicalProperties.line.solidFill = "059669"
    s_vs.graphicalProperties.line.width = 25000
    chart3.series.append(s_vs)

    ws_charts.add_chart(chart3, "B50")

    # Salva o arquivo Excel
    abs_excel_path = os.path.abspath(excel_path)
    wb.save(abs_excel_path)
    print(f"[OK] Planilha Excel profissional salva com sucesso em: {abs_excel_path}")

    # Também salva no nome sem '_excel' para compatibilidade se diferente
    if excel_path != "formas_de_onda_fonte.xlsx":
        try:
            wb.save(os.path.abspath("formas_de_onda_fonte.xlsx"))
        except Exception:
            pass

    return abs_excel_path


# =============================================================================
# 4. VALIDAÇÃO AUTOMATIZADA COM MICROSOFT EXCEL (OPCIONAL VIA COM)
# =============================================================================

def validate_and_refresh_excel(excel_path: str) -> bool:
    """
    Se o Microsoft Excel estiver instalado no Windows, abre a planilha em background
    usando win32com, recalcula todas as fórmulas dinâmicas e salva.
    """
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        abs_path = os.path.abspath(excel_path)
        book = excel.Workbooks.Open(abs_path)
        book.Save()
        book.Close(False)
        excel.Quit()
        print(f"[OK] Fórmulas e gráficos validados com sucesso no motor nativo do Microsoft Excel.")
        return True
    except Exception as e:
        print(f"[INFO] Validação COM do Excel não executada ou dispensada ({e}). Planilha salva em formato padrão OpenXML.")
        return False


# =============================================================================
# 5. BLOCO PRINCIPAL DE EXECUÇÃO
# =============================================================================

if __name__ == "__main__":
    if sys.stdout.encoding != "utf-8":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    params = CircuitParams()
    print("=" * 70)
    print("  RELATÓRIO DE CÁLCULO FÍSICO DO CIRCUITO DE FONTE REGULADA")
    print("=" * 70)
    print(f"  • Frequência da Rede (f):         {params.f_rede:.1f} Hz (T_rede = {params.t_rede*1e3:.2f} ms)")
    print(f"  • Tensão de Entrada Primário:    {params.v_in_ef:.1f} Vef")
    print(f"  • Secundário Transformador:       {params.e0_ef:.3f} Vef (E0_pk = {params.e0_pk:.3f} V)")
    print(f"  • Queda nos Diodos (2x VD):       {2.0 * params.v_diode:.2f} V (2 x {params.v_diode:.1f} V)")
    print(f"  • Nó A - Tensão Máxima (VA_max):  {params.va_max:.2f} V")
    print(f"  • Nó A - Tensão Média (VA_dc):    {params.va_dc:.2f} V")
    print(f"  • Nó A - Tensão Mínima (VA_min):  {params.va_min:.2f} V")
    print(f"  • Nó A - Ripple Pico a Pico:      {params.v_app:.3f} V @ {params.f_ripple:.0f} Hz")
    print(f"  • Nó B - Tensão Regulada (Vs=Vz): {params.v_zener:.2f} V (Diodo 1N4735)")
    print("=" * 70)

    # 1. Gera e salva o gráfico técnico de alta resolução (350 DPI)
    img_target = "formas_de_onda_fonte_excel.png"
    plot_circuit_waveforms(params, save_path=img_target, show_plot=False)

    # 2. Gera a planilha Excel profissional completa com fórmulas, dados e gráficos nativos
    excel_target = "formas_de_onda_fonte_excel.xlsx"
    saved_excel = generate_excel_workbook(
        params,
        excel_path=excel_target,
        img_path=img_target,
        num_excel_points=1000
    )

    # 3. Validação nativa via Excel COM (se disponível)
    validate_and_refresh_excel(saved_excel)

    print("=" * 70)
    print("  PROCESSO CONCLUÍDO COM SUCESSO!")
    print(f"  1. Imagem Técnica:   {os.path.abspath(img_target)}")
    print(f"  2. Planilha Excel:    {saved_excel}")
    print("=" * 70)
